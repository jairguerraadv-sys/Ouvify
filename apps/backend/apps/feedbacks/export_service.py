"""
Export/Import Services - Ouvify
Sprint 5 - Feature 5.3: Export/Import de Dados

Serviços para exportação e importação de dados do sistema.
"""
import csv
import json
import io
import logging
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from django.http import HttpResponse
from django.utils import timezone
from django.db import transaction

from apps.feedbacks.models import Feedback, Tag
from apps.tenants.models import Client

logger = logging.getLogger(__name__)


class ExportService:
    """Serviço de exportação de dados."""
    
    EXPORT_FORMATS = ['csv', 'json', 'xlsx']
    
    @staticmethod
    def export_feedbacks(
        tenant: Client,
        format: str = 'csv',
        filters: Optional[Dict[str, Any]] = None
    ) -> HttpResponse:
        """
        Exporta feedbacks para o formato especificado.
        
        Args:
            tenant: Cliente/tenant
            format: Formato de exportação (csv, json, xlsx)
            filters: Filtros opcionais
        
        Returns:
            HttpResponse com o arquivo
        """
        from apps.feedbacks.models import Feedback
        
        # Usar all_tenants() para bypassa o filtro automático do TenantAwareManager
        queryset = Feedback.objects.all_tenants().filter(client=tenant)
        
        # Aplicar filtros
        if filters:
            if filters.get('periodo'):
                periodo = filters['periodo']
                if periodo != 'all':
                    try:
                        dias = int(periodo)
                        data_inicio = timezone.now() - timedelta(days=dias)
                        queryset = queryset.filter(data_criacao__gte=data_inicio)
                    except ValueError:
                        pass
            
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            
            if filters.get('tipo'):
                queryset = queryset.filter(tipo=filters['tipo'])
            
            if filters.get('prioridade'):
                queryset = queryset.filter(prioridade=filters['prioridade'])
            
            if filters.get('data_inicio'):
                queryset = queryset.filter(data_criacao__gte=filters['data_inicio'])
            
            if filters.get('data_fim'):
                queryset = queryset.filter(data_criacao__lte=filters['data_fim'])
        
        # Ordenar por data de criação
        queryset = queryset.order_by('-data_criacao')
        
        # Exportar no formato solicitado
        if format == 'csv':
            return ExportService._export_feedbacks_csv(queryset, tenant)
        elif format == 'json':
            return ExportService._export_feedbacks_json(queryset, tenant)
        elif format == 'xlsx':
            return ExportService._export_feedbacks_xlsx(queryset, tenant)
        else:
            raise ValueError(f"Formato não suportado: {format}")
    
    @staticmethod
    def _export_feedbacks_csv(queryset, tenant: Client) -> HttpResponse:
        """Exporta feedbacks para CSV."""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"feedbacks_{tenant.subdominio}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # BOM para Excel reconhecer UTF-8
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Protocolo', 'Tipo', 'Título', 'Descrição', 'Status', 'Prioridade',
            'Anônimo', 'Data Criação', 'Data Resolução', 'Tempo Resposta (h)',
            'Tempo Resolução (h)', 'SLA Resposta Cumprido', 'SLA Resolução Cumprido',
            'Atribuído Para', 'Tags', 'Resposta Empresa'
        ])
        
        for fb in queryset.select_related('assigned_to__user').prefetch_related('tags'):
            tempo_resposta = ""
            if fb.tempo_primeira_resposta:
                tempo_resposta = f"{fb.tempo_primeira_resposta.total_seconds() / 3600:.1f}"
            
            tempo_resolucao = ""
            if fb.tempo_resolucao:
                tempo_resolucao = f"{fb.tempo_resolucao.total_seconds() / 3600:.1f}"
            
            assigned_name = ""
            if fb.assigned_to and fb.assigned_to.user:
                assigned_name = fb.assigned_to.user.get_full_name() or fb.assigned_to.user.username
            
            tags = ", ".join([t.nome for t in fb.tags.all()])
            
            writer.writerow([
                fb.id,
                fb.protocolo,
                fb.get_tipo_display() if hasattr(fb, 'get_tipo_display') else fb.tipo,
                fb.titulo,
                fb.descricao[:500] if fb.descricao else '',
                fb.get_status_display() if hasattr(fb, 'get_status_display') else fb.status,
                fb.get_prioridade_display() if hasattr(fb, 'get_prioridade_display') else fb.prioridade,
                'Sim' if fb.anonimo else 'Não',
                fb.data_criacao.strftime('%Y-%m-%d %H:%M'),
                fb.data_resolucao.strftime('%Y-%m-%d %H:%M') if fb.data_resolucao else '',
                tempo_resposta,
                tempo_resolucao,
                'Sim' if fb.sla_primeira_resposta else 'Não' if fb.sla_primeira_resposta is not None else '',
                'Sim' if fb.sla_resolucao else 'Não' if fb.sla_resolucao is not None else '',
                assigned_name,
                tags,
                fb.resposta_empresa[:500] if fb.resposta_empresa else ''
            ])
        
        logger.info(f"📤 CSV exportado | Tenant: {tenant.nome} | Registros: {queryset.count()}")
        return response
    
    @staticmethod
    def _export_feedbacks_json(queryset, tenant: Client) -> HttpResponse:
        """Exporta feedbacks para JSON."""
        data = {
            'meta': {
                'tenant': tenant.nome,
                'exported_at': timezone.now().isoformat(),
                'total_records': queryset.count(),
            },
            'feedbacks': []
        }
        
        for fb in queryset.select_related('assigned_to__user').prefetch_related('tags'):
            assigned_name = ""
            if fb.assigned_to and fb.assigned_to.user:
                assigned_name = fb.assigned_to.user.get_full_name() or fb.assigned_to.user.username
            
            feedback_data = {
                'id': fb.id,
                'protocolo': fb.protocolo,
                'tipo': fb.tipo,
                'tipo_display': fb.get_tipo_display() if hasattr(fb, 'get_tipo_display') else fb.tipo,
                'titulo': fb.titulo,
                'descricao': fb.descricao,
                'status': fb.status,
                'status_display': fb.get_status_display() if hasattr(fb, 'get_status_display') else fb.status,
                'prioridade': fb.prioridade,
                'prioridade_display': fb.get_prioridade_display() if hasattr(fb, 'get_prioridade_display') else fb.prioridade,
                'anonimo': fb.anonimo,
                'data_criacao': fb.data_criacao.isoformat(),
                'data_resolucao': fb.data_resolucao.isoformat() if fb.data_resolucao else None,
                'tempo_primeira_resposta_horas': fb.tempo_primeira_resposta.total_seconds() / 3600 if fb.tempo_primeira_resposta else None,
                'tempo_resolucao_horas': fb.tempo_resolucao.total_seconds() / 3600 if fb.tempo_resolucao else None,
                'sla_primeira_resposta': fb.sla_primeira_resposta,
                'sla_resolucao': fb.sla_resolucao,
                'assigned_to': assigned_name,
                'tags': [t.nome for t in fb.tags.all()],
                'resposta_empresa': fb.resposta_empresa,
            }
            data['feedbacks'].append(feedback_data)
        
        response = HttpResponse(
            json.dumps(data, ensure_ascii=False, indent=2),
            content_type='application/json; charset=utf-8'
        )
        filename = f"feedbacks_{tenant.subdominio}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"📤 JSON exportado | Tenant: {tenant.nome} | Registros: {queryset.count()}")
        return response
    
    @staticmethod
    def _export_feedbacks_xlsx(queryset, tenant: Client) -> HttpResponse:
        """Exporta feedbacks para Excel (XLSX)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            # Se openpyxl não estiver instalado, retorna CSV
            logger.warning("openpyxl não instalado, exportando como CSV")
            return ExportService._export_feedbacks_csv(queryset, tenant)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Feedbacks"
        
        # Cabeçalho com estilo
        headers = [
            'ID', 'Protocolo', 'Tipo', 'Título', 'Descrição', 'Status', 'Prioridade',
            'Anônimo', 'Data Criação', 'Data Resolução', 'Tempo Resposta (h)',
            'Tempo Resolução (h)', 'SLA Resposta', 'SLA Resolução',
            'Atribuído Para', 'Tags', 'Resposta Empresa'
        ]
        
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        for row, fb in enumerate(queryset.select_related('assigned_to__user').prefetch_related('tags'), 2):
            tempo_resposta = None
            if fb.tempo_primeira_resposta:
                tempo_resposta = round(fb.tempo_primeira_resposta.total_seconds() / 3600, 1)
            
            tempo_resolucao = None
            if fb.tempo_resolucao:
                tempo_resolucao = round(fb.tempo_resolucao.total_seconds() / 3600, 1)
            
            assigned_name = ""
            if fb.assigned_to and fb.assigned_to.user:
                assigned_name = fb.assigned_to.user.get_full_name() or fb.assigned_to.user.username
            
            tags = ", ".join([t.nome for t in fb.tags.all()])
            
            ws.cell(row=row, column=1, value=fb.id)
            ws.cell(row=row, column=2, value=fb.protocolo)
            ws.cell(row=row, column=3, value=fb.get_tipo_display() if hasattr(fb, 'get_tipo_display') else fb.tipo)
            ws.cell(row=row, column=4, value=fb.titulo)
            ws.cell(row=row, column=5, value=fb.descricao[:500] if fb.descricao else '')
            ws.cell(row=row, column=6, value=fb.get_status_display() if hasattr(fb, 'get_status_display') else fb.status)
            ws.cell(row=row, column=7, value=fb.get_prioridade_display() if hasattr(fb, 'get_prioridade_display') else fb.prioridade)
            ws.cell(row=row, column=8, value='Sim' if fb.anonimo else 'Não')
            ws.cell(row=row, column=9, value=fb.data_criacao.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=10, value=fb.data_resolucao.strftime('%Y-%m-%d %H:%M') if fb.data_resolucao else '')
            ws.cell(row=row, column=11, value=tempo_resposta)
            ws.cell(row=row, column=12, value=tempo_resolucao)
            ws.cell(row=row, column=13, value='Sim' if fb.sla_primeira_resposta else 'Não' if fb.sla_primeira_resposta is not None else '')
            ws.cell(row=row, column=14, value='Sim' if fb.sla_resolucao else 'Não' if fb.sla_resolucao is not None else '')
            ws.cell(row=row, column=15, value=assigned_name)
            ws.cell(row=row, column=16, value=tags)
            ws.cell(row=row, column=17, value=fb.resposta_empresa[:500] if fb.resposta_empresa else '')
        
        # Auto-ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Salvar em buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"feedbacks_{tenant.subdominio}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"📤 XLSX exportado | Tenant: {tenant.nome} | Registros: {queryset.count()}")
        return response


class ImportService:
    """Serviço de importação de dados."""
    
    IMPORT_FORMATS = ['csv', 'json']
    
    @staticmethod
    @transaction.atomic
    def import_feedbacks(
        tenant: Client,
        file_content: bytes,
        format: str = 'csv',
        update_existing: bool = False
    ) -> Dict[str, Any]:
        """
        Importa feedbacks de um arquivo.
        
        Args:
            tenant: Cliente/tenant
            file_content: Conteúdo do arquivo
            format: Formato do arquivo (csv, json)
            update_existing: Se True, atualiza feedbacks existentes
        
        Returns:
            Resultado da importação
        """
        if format == 'csv':
            return ImportService._import_feedbacks_csv(tenant, file_content, update_existing)
        elif format == 'json':
            return ImportService._import_feedbacks_json(tenant, file_content, update_existing)
        else:
            raise ValueError(f"Formato não suportado: {format}")
    
    @staticmethod
    def _import_feedbacks_csv(
        tenant: Client,
        file_content: bytes,
        update_existing: bool
    ) -> Dict[str, Any]:
        """Importa feedbacks de CSV."""
        result = {
            'success': True,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': []
        }
        
        try:
            content = file_content.decode('utf-8-sig')  # Remove BOM se existir
            reader = csv.DictReader(io.StringIO(content))
            
            for row_num, row in enumerate(reader, 2):
                try:
                    protocolo = row.get('Protocolo', '').strip()
                    
                    # Verificar se já existe
                    existing = Feedback.objects.all_tenants().filter(
                        client=tenant,
                        protocolo=protocolo
                    ).first() if protocolo else None
                    
                    if existing:
                        if update_existing:
                            # Atualizar feedback existente
                            existing.titulo = row.get('Título', existing.titulo)
                            existing.descricao = row.get('Descrição', existing.descricao)
                            existing.save()
                            result['updated'] += 1
                        else:
                            result['skipped'] += 1
                    else:
                        # Criar novo feedback
                        tipo = ImportService._normalize_tipo(row.get('Tipo', 'sugestao'))
                        status = ImportService._normalize_status(row.get('Status', 'aberto'))
                        prioridade = ImportService._normalize_prioridade(row.get('Prioridade', 'media'))
                        
                        Feedback.objects.create(
                            client=tenant,
                            tipo=tipo,
                            titulo=row.get('Título', 'Feedback Importado'),
                            descricao=row.get('Descrição', ''),
                            status=status,
                            prioridade=prioridade,
                            anonimo=row.get('Anônimo', '').lower() in ['sim', 'true', '1'],
                        )
                        result['created'] += 1
                        
                except Exception as e:
                    result['errors'].append(f"Linha {row_num}: {str(e)}")
                    
        except Exception as e:
            result['success'] = False
            result['errors'].append(f"Erro ao processar arquivo: {str(e)}")
        
        logger.info(
            f"📥 CSV importado | Tenant: {tenant.nome} | "
            f"Criados: {result['created']} | Atualizados: {result['updated']} | "
            f"Ignorados: {result['skipped']} | Erros: {len(result['errors'])}"
        )
        return result
    
    @staticmethod
    def _import_feedbacks_json(
        tenant: Client,
        file_content: bytes,
        update_existing: bool
    ) -> Dict[str, Any]:
        """Importa feedbacks de JSON."""
        result = {
            'success': True,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': []
        }
        
        try:
            data = json.loads(file_content.decode('utf-8'))
            feedbacks = data.get('feedbacks', data if isinstance(data, list) else [])
            
            for idx, fb_data in enumerate(feedbacks):
                try:
                    protocolo = fb_data.get('protocolo', '').strip()
                    
                    existing = Feedback.objects.all_tenants().filter(
                        client=tenant,
                        protocolo=protocolo
                    ).first() if protocolo else None
                    
                    if existing:
                        if update_existing:
                            existing.titulo = fb_data.get('titulo', existing.titulo)
                            existing.descricao = fb_data.get('descricao', existing.descricao)
                            existing.save()
                            result['updated'] += 1
                        else:
                            result['skipped'] += 1
                    else:
                        Feedback.objects.create(
                            client=tenant,
                            tipo=ImportService._normalize_tipo(fb_data.get('tipo', 'sugestao')),
                            titulo=fb_data.get('titulo', 'Feedback Importado'),
                            descricao=fb_data.get('descricao', ''),
                            status=ImportService._normalize_status(fb_data.get('status', 'aberto')),
                            prioridade=ImportService._normalize_prioridade(fb_data.get('prioridade', 'media')),
                            anonimo=fb_data.get('anonimo', False),
                        )
                        result['created'] += 1
                        
                except Exception as e:
                    result['errors'].append(f"Item {idx + 1}: {str(e)}")
                    
        except json.JSONDecodeError as e:
            result['success'] = False
            result['errors'].append(f"JSON inválido: {str(e)}")
        except Exception as e:
            result['success'] = False
            result['errors'].append(f"Erro ao processar arquivo: {str(e)}")
        
        logger.info(
            f"📥 JSON importado | Tenant: {tenant.nome} | "
            f"Criados: {result['created']} | Atualizados: {result['updated']} | "
            f"Ignorados: {result['skipped']} | Erros: {len(result['errors'])}"
        )
        return result
    
    @staticmethod
    def _normalize_tipo(tipo: str) -> str:
        """Normaliza o tipo de feedback."""
        tipo_map = {
            'reclamação': 'reclamacao',
            'reclamacao': 'reclamacao',
            'sugestão': 'sugestao',
            'sugestao': 'sugestao',
            'elogio': 'elogio',
            'dúvida': 'duvida',
            'duvida': 'duvida',
            'solicitação': 'solicitacao',
            'solicitacao': 'solicitacao',
        }
        return tipo_map.get(tipo.lower(), 'sugestao')
    
    @staticmethod
    def _normalize_status(status: str) -> str:
        """Normaliza o status."""
        status_map = {
            'aberto': 'aberto',
            'em análise': 'em_analise',
            'em análise': 'em_analise',
            'em_analise': 'em_analise',
            'em andamento': 'em_andamento',
            'em_andamento': 'em_andamento',
            'aguardando': 'aguardando',
            'resolvido': 'resolvido',
            'fechado': 'fechado',
        }
        return status_map.get(status.lower(), 'aberto')
    
    @staticmethod
    def _normalize_prioridade(prioridade: str) -> str:
        """Normaliza a prioridade."""
        prioridade_map = {
            'baixa': 'baixa',
            'média': 'media',
            'media': 'media',
            'alta': 'alta',
            'urgente': 'urgente',
            'crítica': 'critica',
            'critica': 'critica',
        }
        return prioridade_map.get(prioridade.lower(), 'media')
